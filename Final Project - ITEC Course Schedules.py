''''' Pre Project Notes

Write a program that reads the MCTC "Course Search Results" page, parses that page and creates a spreadsheet containing 7 columns
1. ID Number (ie. 000096)
2. Course Number (ie. 1150-61)
3. Course Title (ie. Programming Logic and Design)
4. Day of the week (ie. Wednesday)
5. Time the course meet (ie. 9:00am - 11:59am)
6. The credits for the course
7. The instructor'''

# requests module lets you easily download files from the Web
import requests
# Had to install beautifulsoup4 or bs4 which allows me to extract information from an HTML page
from bs4 import BeautifulSoup
# Allows me to write to Excel
from openpyxl import Workbook


# Request the URL we want to work with
res = requests.get('https://eservices.minnstate.edu/registration/search/advancedSubmit.html?campusid=305&searchrcid'
                   '=0305&searchcampusid=305&yrtr=20205&subject=ITEC&courseNumber=&courseId=&openValue'
                   '=OPEN_PLUS_WAITLIST&delivery=ALL&showAdvanced=&starttime=&endtime=&mntransfer=&credittype=ALL'
                   '&credits=&instructor=&keyword=&begindate=&site=&resultNumber=250')
try:
    # raise_for_status raises an exception if there was an error downloading the file and will do nothing if the
    # download succeeded
    res.raise_for_status()
except Exception as exc:
    # prints this statement if download failed "There was a problem: 404 Client Error: Not Found"
    print('There was a problem: %s' % (exc))

# Creates a beautiful soup object that is stored in the variable soup. Gives us the HTML of page
soup = BeautifulSoup(res.content, 'html.parser')

# Grab the table of web page and use the data within it. Inside of <div id="resultsContainer">
tableTag = soup.find(id='resultsTable')

# Create a list where row data will go into
resRows = []
# Finds all the 'tr' tags which defines a row in the table
rowTags = tableTag.find_all('tr')
# Use the first list where we find all the 'th' tags which defines a header cell in a table. This is how we start to
# break rows into columns
headerTags = rowTags[0].findAll('th')
# List Created where the headers found will go into and locate need columns
headers = []
# For all headers found in headerTags
for headerTag in headerTags:
    # Add just the text to headers list and remove leading and trailing spaces with strip()
    headers.append(headerTag.text.strip())
# Print statement below used to test if we got the headers
# print(headers)

# For the sections in the rows
for rowTag in rowTags:
    # Extract the 'td tags from every row
    columnTags = rowTag.findAll('td')
    # Create list of rows split by columns
    resCol = []
    # For section in rows
    for columnTag in columnTags:
        # Add list of just the text to resCol list
        resCol.append(columnTag.text.strip())
    # Add new row sectioned list to master list resRows
    resRows.append(resCol)
# Print statement below used to test if we get lists of rows
# print(resRows)

# 1-'ID #', 3-'#', 5-'Title', 7-'Days', 8-'Time', 9-'Cr/Hr', 11-'Instructor'
colIndices = [1, 3, 5, 7, 8, 9, 11]
# List of list column text will be stored in
listCols = [[], [], [], [], [], [], []]
# For sections in second row list
for rowTag in rowTags[1:]:
    # Find all cells
    columnTags = rowTag.findAll('td')
    # Integer created that starts at 0
    i = 0
    # For certain list within colIndices
    for colIndex in colIndices:
        # Get just the text and remove certain text and lines with spaces instead
        value = columnTags[colIndex].text.strip().replace('\xa0', ' ').replace('\n', ' ')
        # Start list with assign number 0 and add new edited list without extra text
        listCols[i].append(value)
        # Designates which column goes with what list starting with 0 increasing by one
        i = i + 1
# print statement below used to test if we get just a clean version of the list of just text
# print(listCols)

# Makes a new workbook
workbook = Workbook()

# Get the active sheet
worksheet = workbook.active

# Writes a column title to a cell row = 1, column = 1
worksheet.cell(1, 1, 'ID Number')
# row = 1, column = 2
worksheet.cell(1, 2, 'Course Number')
# row = 1, column = 3 etc
worksheet.cell(1, 3, 'Course Title')
worksheet.cell(1, 4, 'Day')
worksheet.cell(1, 5, 'Time')
worksheet.cell(1, 6, 'Credits')
worksheet.cell(1, 7, 'Instructor')

# In the list of list, listCols it pulls the first list and prints to row = 2, column = 1
for index, id_numbers in enumerate(listCols[0]):
    worksheet.cell(index + 2, 1, id_numbers)
# prints course numbers starting in row 2, column 2 and down the rows
for index, course_number in enumerate(listCols[1]):
    worksheet.cell(index + 2, 2, course_number)
# prints course titles starting in row 2, column 3 and down the rows
for index, course_title in enumerate(listCols[2]):
    worksheet.cell(index + 2, 3, course_title)
# prints days of the week starting in row 2, column 4 and down the rows
for index, days in enumerate(listCols[3]):
    worksheet.cell(index + 2, 4, days)
# prints times classes meet starting in row 2, column 5 and down the rows
for index, times in enumerate(listCols[4]):
    worksheet.cell(index + 2, 5, times)
# prints the number of credits starting in row 2, column 6 and down the rows
for index, numb_of_credit in enumerate(listCols[5]):
    worksheet.cell(index + 2, 6, numb_of_credit)
# prints the 7th list in listCols starting in row 2, column 7 which are the instructors names
for index, instructors in enumerate(listCols[6]):
    worksheet.cell(index + 2, 7, instructors)

# Saves workbook
workbook.save('Final Project - ITEC Course Schedules.xlsx')

# Github URL: https://github.com/johnnyvuanly/personal/blob/master/Final%20Project%20-%20ITEC%20Course%20Schedules.py